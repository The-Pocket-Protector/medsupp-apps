#!/usr/bin/env python3
"""
Build Excel carrier report from SERFF form filing data.
Report: State | Insurance Carrier | # Approved Form Filings
"""

import json
from pathlib import Path
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT_DIR = Path("/home/openclaw/.openclaw/workspace/medsupp-apps/output/serff")
EXCEL_PATH = Path("/home/openclaw/.openclaw/workspace/medsupp-apps/output/medsupp_carrier_report.xlsx")


def is_approved(status):
    s = (status or "").lower()
    return "approved" in s and "disapproved" not in s


def build_report():
    files = sorted(OUTPUT_DIR.glob("*_form_filings.json"))
    print(f"Loading {len(files)} state files...")

    # Collect: state -> set of approved carriers
    state_carriers = defaultdict(set)
    state_filing_counts = defaultdict(lambda: defaultdict(int))
    state_totals = {}
    all_carriers = set()

    for f in files:
        data = json.loads(f.read_text())
        state = data["state"]
        state_totals[state] = data.get("total_in_serff", 0)

        for row in data.get("form_rows", []):
            if is_approved(row.get("Filing Status")):
                carrier = row.get("Company Name", "Unknown").strip()
                if carrier and carrier != "Unknown":
                    state_carriers[state].add(carrier)
                    state_filing_counts[state][carrier] += 1
                    all_carriers.add(carrier)

    print(f"States with data: {len(state_carriers)}")
    print(f"Unique carriers across all states: {len(all_carriers)}")

    # Create workbook
    wb = openpyxl.Workbook()

    # ── Sheet 1: Summary by State ─────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "By State"

    # Header style
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    alt_fill = PatternFill(start_color="EBF3FB", end_color="EBF3FB", fill_type="solid")
    border = Border(
        bottom=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC")
    )

    # Headers
    ws1["A1"] = "State"
    ws1["B1"] = "Insurance Carrier"
    ws1["C1"] = "# Approved Form Filings"
    ws1["D1"] = "Total MS Filings in SERFF"

    for cell in ws1[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    ws1.row_dimensions[1].height = 30

    # Data rows
    row_idx = 2
    sorted_states = sorted(state_carriers.keys())

    for state in sorted_states:
        carriers = sorted(state_carriers[state])
        total_ms = state_totals.get(state, 0)

        for i, carrier in enumerate(carriers):
            fill = alt_fill if row_idx % 2 == 0 else PatternFill(fill_type=None)
            count = state_filing_counts[state][carrier]

            ws1.cell(row=row_idx, column=1, value=state if i == 0 else "")
            ws1.cell(row=row_idx, column=2, value=carrier)
            ws1.cell(row=row_idx, column=3, value=count)
            ws1.cell(row=row_idx, column=4, value=total_ms if i == 0 else "")

            for col in range(1, 5):
                cell = ws1.cell(row=row_idx, column=col)
                if row_idx % 2 == 0:
                    cell.fill = alt_fill
                cell.border = border
                cell.alignment = Alignment(vertical="center")

            row_idx += 1

        # Add state subtotal row
        subtotal_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
        subtotal_font = Font(bold=True, italic=True)
        ws1.cell(row=row_idx, column=1, value=f"{state} TOTAL")
        ws1.cell(row=row_idx, column=2, value=f"{len(carriers)} carriers")
        ws1.cell(row=row_idx, column=3, value=sum(state_filing_counts[state].values()))
        ws1.cell(row=row_idx, column=4, value="")

        for col in range(1, 5):
            cell = ws1.cell(row=row_idx, column=col)
            cell.fill = subtotal_fill
            cell.font = subtotal_font
            cell.border = border
            cell.alignment = Alignment(vertical="center")

        row_idx += 1

    # Column widths
    ws1.column_dimensions["A"].width = 8
    ws1.column_dimensions["B"].width = 50
    ws1.column_dimensions["C"].width = 22
    ws1.column_dimensions["D"].width = 22

    # Freeze header row
    ws1.freeze_panes = "A2"

    # ── Sheet 2: Carrier Pivot ────────────────────────────────────────────
    ws2 = wb.create_sheet("Carrier by State")

    sorted_carriers = sorted(all_carriers)
    sorted_states_list = sorted(state_carriers.keys())

    # Headers
    ws2["A1"] = "Insurance Carrier"
    for j, state in enumerate(sorted_states_list):
        ws2.cell(row=1, column=j + 2, value=state)

    for cell in ws2[1]:
        if cell.value:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

    ws2.row_dimensions[1].height = 25

    # Carrier rows
    for i, carrier in enumerate(sorted_carriers, start=2):
        ws2.cell(row=i, column=1, value=carrier)
        fill = alt_fill if i % 2 == 0 else PatternFill(fill_type=None)
        ws2.cell(row=i, column=1).fill = fill if i % 2 == 0 else PatternFill(fill_type=None)

        for j, state in enumerate(sorted_states_list):
            count = state_filing_counts[state].get(carrier, 0)
            cell = ws2.cell(row=i, column=j + 2, value=count if count > 0 else "")
            if i % 2 == 0:
                cell.fill = alt_fill
            if count > 0:
                cell.alignment = Alignment(horizontal="center")
            cell.border = border

        ws2.cell(row=i, column=1).border = border
        ws2.cell(row=i, column=1).alignment = Alignment(vertical="center")

    ws2.column_dimensions["A"].width = 50
    for j in range(len(sorted_states_list)):
        ws2.column_dimensions[get_column_letter(j + 2)].width = 7

    ws2.freeze_panes = "B2"

    # ── Sheet 3: Summary Overview ─────────────────────────────────────────
    ws3 = wb.create_sheet("Summary")

    ws3["A1"] = "SERFF Medicare Supplement - Approved Form Filings"
    ws3["A1"].font = Font(bold=True, size=14, color="1F4E79")
    ws3.merge_cells("A1:D1")

    ws3["A3"] = "State"
    ws3["B3"] = "Total Approved Carriers"
    ws3["C3"] = "Total Approved Filings"
    ws3["D3"] = "Total MS Filings in SERFF"

    for cell in ws3[3]:
        if cell.value:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

    grand_carriers = 0
    grand_filings = 0

    for i, state in enumerate(sorted_states, start=4):
        carriers = state_carriers[state]
        total_approved = sum(state_filing_counts[state].values())
        fill = alt_fill if i % 2 == 0 else PatternFill(fill_type=None)

        ws3.cell(row=i, column=1, value=state)
        ws3.cell(row=i, column=2, value=len(carriers))
        ws3.cell(row=i, column=3, value=total_approved)
        ws3.cell(row=i, column=4, value=state_totals.get(state, 0))

        for col in range(1, 5):
            cell = ws3.cell(row=i, column=col)
            if i % 2 == 0:
                cell.fill = alt_fill
            cell.alignment = Alignment(horizontal="center" if col > 1 else "left", vertical="center")
            cell.border = border

        grand_carriers += len(carriers)
        grand_filings += total_approved

    # Grand total
    total_row = len(sorted_states) + 4
    ws3.cell(row=total_row, column=1, value="TOTAL")
    ws3.cell(row=total_row, column=2, value=len(all_carriers))
    ws3.cell(row=total_row, column=3, value=grand_filings)
    ws3.cell(row=total_row, column=4, value=sum(state_totals.values()))

    total_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    total_font = Font(bold=True, color="FFFFFF")
    for col in range(1, 5):
        cell = ws3.cell(row=total_row, column=col)
        cell.fill = total_fill
        cell.font = total_font
        cell.alignment = Alignment(horizontal="center" if col > 1 else "left", vertical="center")

    ws3.column_dimensions["A"].width = 8
    ws3.column_dimensions["B"].width = 22
    ws3.column_dimensions["C"].width = 22
    ws3.column_dimensions["D"].width = 24
    ws3.freeze_panes = "A4"

    # Save
    wb.save(EXCEL_PATH)
    print(f"\nSaved: {EXCEL_PATH}")
    print(f"Sheets: {[ws.title for ws in wb.worksheets]}")
    print(f"\nSummary:")
    print(f"  States covered: {len(sorted_states)}")
    print(f"  Unique carriers: {len(all_carriers)}")
    print(f"  Total approved filings: {grand_filings}")


if __name__ == "__main__":
    build_report()
